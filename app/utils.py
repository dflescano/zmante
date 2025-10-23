from datetime import datetime, timedelta
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO

## utils.py — helper

from datetime import datetime, date, timedelta

def pc_created_date(pc):
    from .models import ChangeLog, Maintenance, Backup
    cl = (ChangeLog.query
          .filter_by(entity="PC", entity_id=pc.id, action="create")
          .order_by(ChangeLog.created_at.asc())
          .first())
    if cl and cl.created_at:
        try:
            return cl.created_at.date()
        except Exception:
            pass
    if hasattr(pc, "created_at") and pc.created_at:
        try:
            return pc.created_at.date() if hasattr(pc.created_at, "date") else pc.created_at
        except Exception:
            pass
    first_m = (Maintenance.query.filter_by(pc_id=pc.id)
               .order_by(Maintenance.date_performed.asc()).first())
    first_b = (Backup.query.filter_by(pc_id=pc.id)
               .order_by(Backup.date_performed.asc()).first())
    candidates = []
    if first_m and first_m.date_performed:
        try: candidates.append(first_m.date_performed.date())
        except Exception: pass
    if first_b and first_b.date_performed:
        try: candidates.append(first_b.date_performed.date())
        except Exception: pass
    if candidates:
        return min(candidates)
    return date.today() - timedelta(days=365)
## utils.py — compute_status
def compute_status(pc, maint_days=7):
    today = datetime.now().date()
    last = pc.last_maintenance_date()
    if last is None:
        start = pc_created_date(pc)
        delta = (today - start).days
        if delta > maint_days:
            return f"SIN MANTENIMIENTO (ALERTA, {delta} días)"
        elif maint_days - 2 <= delta <= maint_days:
            return f"SIN MANTENIMIENTO (POR VENCER, {delta} días)"
        else:
            return f"SIN MANTENIMIENTO ({delta} días)"
    delta = (today - last).days
    if delta > maint_days:
        return f"ALERTA ({delta} días)"
    elif maint_days - 2 <= delta <= maint_days:
        return f"POR VENCER ({delta} días)"
    else:
        return f"OK ({delta} días)"

def pcs_to_workbook(pcs, maint_days=7):
    wb = Workbook(); ws = wb.active; ws.title = "PCs"
    ws.append(["PC","Usuario PC","Usuario físico","TeamViewer","AnyDesk","Windows legal","Office legal","Ubicación","Último mant.","Estado"])
    for pc in pcs:
        last = pc.last_maintenance().date_performed.strftime("%Y-%m-%d %H:%M") if pc.last_maintenance() else "—"
        ws.append([pc.name, pc.pc_username or "", pc.physical_user or "", pc.teamviewer_id or "", pc.anydesk_id or "",
                   "Sí" if pc.windows_licensed else "No", "Sí" if pc.office_licensed else "No",
                   pc.location or "", last, compute_status(pc, maint_days)])
    return wb

def activity_to_workbook(maintenances, backups):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Mantenimientos"
    ws1.append(["PC","Fecha","Técnico","Detalle"])
    for m in maintenances:
        ws1.append([m.pc.name, m.date_performed.strftime("%Y-%m-%d %H:%M"), m.performed_by or "", m.description or ""])
    ws2 = wb.create_sheet("Backups")
    ws2.append(["PC","Fecha","Estado","Tamaño (MB)","Ruta"])
    for b in backups:
        ws2.append([b.pc.name, b.date_performed.strftime("%Y-%m-%d %H:%M"), b.status, b.size_mb or "", b.path or ""])
    return wb

def activity_to_pdf(maintenances, backups):
    bio = BytesIO(); c = canvas.Canvas(bio, pagesize=A4)
    w,h = A4; margin=40; y=h-margin
    c.setFont("Helvetica-Bold", 12); c.drawString(margin,y,"Actividad (Mantenimientos y Backups)"); y-=24
    c.setFont("Helvetica-Bold", 10); c.drawString(margin,y,"Mantenimientos"); y-=16; c.setFont("Helvetica",9)
    for m in maintenances:
        if y<60: c.showPage(); y=h-margin; c.setFont("Helvetica",9)
        c.drawString(margin,y,f"{m.date_performed.strftime('%Y-%m-%d %H:%M')} - {m.pc.name} - {m.performed_by or ''} - {m.description or ''}"); y-=14
    y-=12; c.setFont("Helvetica-Bold", 10); c.drawString(margin,y,"Backups"); y-=16; c.setFont("Helvetica",9)
    for b in backups:
        if y<60: c.showPage(); y=h-margin; c.setFont("Helvetica",9)
        c.drawString(margin,y,f"{b.date_performed.strftime('%Y-%m-%d %H:%M')} - {b.pc.name} - {b.status} - {b.size_mb or ''} MB - {b.path or ''}"); y-=14
    c.showPage(); c.save(); bio.seek(0); return bio
